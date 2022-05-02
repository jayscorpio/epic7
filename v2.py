from openpyxl import load_workbook
import multiprocessing as mp
from configparser import ConfigParser
import os
from hero import Database, Hero
import re
# TBD
# Save priorized item sets to a file for reuse
# Constant
PRINT_ITEMS_KINDS = {
    'weapon': '武器',
    'head': '头盔',
    'armor': '盔甲',
    'neck': '项链',
    'ring': '戒指',
    'shoe': '鞋子'
}
ITEMS_LIST = [
    ['weapon', 0],
    ['head', 1],
    ['armor', 2],
    ['neck', 3],
    ['ring', 4],
    ['shoe', 5]
]
ITEM_SETS_LIST = [
    'AT',
    'DF',
    'HP',
    'SP',
    'CT',
    'CD',
    'HT',
    'EV',
    'FU',
    'TO',
    'PC',
]
BASIC_STATES_LIST = [
    'AT',
    'DF',
    'HP',
    'SP',
    'CT',
    'CD',
    'HT',
    'EV'
]
# THRESHOLDS_LIST = [
#     'CTo',
#     'HTo',
#     'SPo',
#     'TANKo',
#     'EVo'
# ]
# Excel format
PLAN_COL_THRESHOLD_STATE = {
    4: 'CT',
    5: 'HT',
    6: 'SP',
    7: 'TANK',
    8: 'EV'
}
PLAN_COL_THRESHOLD_SET = {
    9: 'BL',
    10: 'FB',
    11: 'IM',
}
PLAN_COL_ARTIFACT = 12
PLAN_COL_STATE = {
    26: 'ATa',
    27: 'ATp',
    28: 'DFa',
    29: 'DFp',
    30: 'HPa',
    31: 'HPp',
    32: 'SPa',
    33: 'CTa',
    34: 'CDa',
    35: 'HTa',
    36: 'HPp',
    37: 'EVa',
}
PLAN_COL_EXTRA = {
    13: 'IgnoreFU',
    14: 'IgnorePC',
    15: 'IgnoreCT',
}
class Sample:
    def __init__(self):
        """Init
        """
        self.items = []
    def copy(self):
        """Make a copy
        Inherits formula, states, and thresholds.
        Inherits NOT items.

        Returns:
            Hero: the copy
        """
        new_hero = Sample()
        new_hero.AT         = self.AT
        new_hero.DF         = self.DF
        new_hero.HP         = self.HP
        new_hero.SP         = self.SP
        new_hero.CT         = self.CT
        new_hero.CD         = self.CD
        new_hero.HT         = self.HT
        new_hero.EV         = self.EV
        new_hero.ATp        = self.ATp
        new_hero.DFp        = self.DFp
        new_hero.HPp        = self.HPp
        new_hero.ATa        = self.ATa
        new_hero.DFa        = self.DFa
        new_hero.HPa        = self.HPa
        new_hero.SPa        = self.SPa
        new_hero.CTa        = self.CTa
        new_hero.CDa        = self.CDa
        new_hero.HTa        = self.HTa
        new_hero.EVa        = self.EVa
        return new_hero
    def load(self, dict_data):
        """Load basic states

        Args:
            dict_data (dict): states data
        """
        self.AT = dict_data['AT']
        self.DF = dict_data['DF']
        self.HP = dict_data['HP']
        self.SP = dict_data['SP']
        self.CT = dict_data['CT']
        self.CD = dict_data['CD']
        self.HT = dict_data['HT']
        self.EV = dict_data['EV']
        self.lv = dict_data['lv']
        self.rank = dict_data['rank']
        self.star = dict_data['star']
    def load_plan(self, dict_data):
        """Load plan
        Including:
            states update
            artifact
        Args:
            dict_data (dict): Plan to be calculated
        """
        self.ATp = float(dict_data['states']['ATp'])
        self.DFp = float(dict_data['states']['DFp'])
        self.HPp = float(dict_data['states']['HPp'])
        self.ATa = int(dict_data['states']['ATa'])
        self.DFa = int(dict_data['states']['DFa'])
        self.HPa = int(dict_data['states']['HPa'])
        self.SPa = int(dict_data['states']['SPa'])
        self.CTa = float(dict_data['states']['CTa'])
        self.CDa = int(dict_data['states']['CDa'])
        self.HTa = float(dict_data['states']['HTa'])
        self.EVa = float(dict_data['states']['EVa'])
        # Artifact
        name = dict_data['artifact']
        if name != None:
            if name not in artifacts:
                print('error artifact [{nm}] not found'.format(nm=name))
                return ('404', name)
            self.ATa += int(artifacts[name]['ATa'])
            self.HPa += int(artifacts[name]['HPa'])
        return (None, None)
    def chk_thresholds(self, thresholds, status=None):
        """To Check thresholds pass or not
        TBD: It would be better to be used externally from Hero class

        Returns:
            bool: pass thresholds check
        """
        """ Only update status when not provided for efficiency """
        if status == None:
            status = self.get_status()
        for criteria, value in thresholds.items():
            if criteria == 'CTo':
                if status['CT'] < value:
                    return False
            elif criteria == 'HTo':
                if status['HT'] < value:
                    return False
            elif criteria == 'SPo':
                if status['SP'] < value:
                    return False
            elif criteria == 'SPu':
                if status['SP'] > value:
                    return False
            elif criteria == 'TANKo':
                if self.calc_criteria('TANK', extra, status) < value:
                    return False
            elif criteria == 'EVo':
                if status['EV'] < value:
                    return False
            else:
                print('error')
        return True
    def strip(self):
        """To remove all clothes
        """
        self.items = []
    def equip(self, item):
        """To equip one item

        Args:
            item (dict): item data
        """
        self.items.append(item)
    def get_sets(self):
        result = set()
        counter = {
            'FB': 0,
            'IM': 0,
            'BL': 0,
        }
        for itm in self.items:
            if itm['set'] not in counter:
                continue
            counter[itm['set']] += 1
        # FB
        if counter['FB'] > 3:
            result.add('FB')
        # IM
        if counter['IM'] > 1:
            result.add('IM')
        # BL
        if counter['BL'] > 3:
            result.add('BL')
        return result
    def get_status(self):
        """Get hero status

        Returns:
            dict: all kinds of status
        """
        counter_sets = {}
        for i in ITEM_SETS_LIST:
            counter_sets[i] = 0
        data = {
            'AT': self.AT,
            'DF': self.DF,
            'HP': self.HP,
            'SP': self.SP,
            'CT': self.CT,
            'CD': self.CD,
            'HT': self.HT,
            'EV': self.EV,
            'ATp': self.ATp,
            'DFp': self.DFp,
            'HPp': self.HPp,
            'ATa': self.ATa,
            'DFa': self.DFa,
            'HPa': self.HPa,
            'SPa': self.SPa,
            'CTa': self.CTa,
            'CDa': self.CDa,
            'HTa': self.HTa,
            'EVa': self.EVa,
        }
        for itm in self.items:
            if itm['set'] in counter_sets:
                counter_sets[itm['set']] += 1
            for attrb in itm['attributes']:
                data[attrb['type']] += attrb['value']
        result = {}
        # AT
        vl = data['AT']
        vl_p = data['ATp']
        vl_a = data['ATa']
        if counter_sets['AT'] > 3:
            vl_p += 45
        result['AT'] = vl * (100 + vl_p) / 100 + vl_a
        # DF
        vl = data['DF']
        vl_p = data['DFp']
        vl_a = data['DFa']
        if counter_sets['DF'] > 1:
            vl_p += 20
            if counter_sets['DF'] > 3:
                vl_p += 20
                if counter_sets['DF'] > 5:
                    vl_p += 20
        result['DF'] = vl * (100 + vl_p) / 100 + vl_a
        # HP
        vl = data['HP']
        vl_p = data['HPp']
        vl_a = data['HPa']
        if counter_sets['HP'] > 1:
            vl_p += 20
            if counter_sets['HP'] > 3:
                vl_p += 20
                if counter_sets['HP'] > 5:
                    vl_p += 20
        result['HP'] = vl * (100 + vl_p) / 100 + vl_a
        # SP
        vl = data['SP']
        vl_a = data['SPa']
        if counter_sets['SP'] > 3:
            vl_p = 25
        else:
            vl_p = 0
        result['SP'] = vl * (100 + vl_p) / 100 + vl_a
        # CT
        vl = data['CT'] + data['CTa']
        if counter_sets['CT'] > 1:
            vl += 12
            if counter_sets['CT'] > 3:
                vl += 12
                if counter_sets['CT'] > 5:
                    vl += 12
        if vl > 100:
            vl = 100
        result['CT'] = vl
        # CD
        vl = data['CD'] + data['CDa']
        if counter_sets['CD'] > 3:
            vl += 40
        if vl > 350:
            vl = 350
        result['CD'] = vl
        # HT
        vl = data['HT'] + data['HTa']
        if counter_sets['HT'] > 1:
            vl += 20
            if counter_sets['HT'] > 3:
                vl += 20
                if counter_sets['HT'] > 5:
                    vl += 20
        result['HT'] = vl
        # EV
        vl = data['EV'] + data['EVa']
        if counter_sets['EV'] > 1:
            vl += 20
            if counter_sets['EV'] > 3:
                vl += 20
                if counter_sets['EV'] > 5:
                    vl += 20
        result['EV'] = vl
        # FU
        result['FU'] = (counter_sets['FU'] > 3)
        # PC
        result['PC'] = (counter_sets['PC'] > 1)
        return result
    def get_benchmark(self, formula, extra, status=None):
        """To calc benchmark

        Returns:
            list: benchmark
        """
        benchmark = []
        """ Only update status when not provided for efficiency """
        if status == None:
            status = self.get_status()
        for crtr in formula:
            benchmark.append(self.calc_criteria(crtr, extra, status))
        return benchmark
    def calc_criteria(self, criteria, extra, status):
        """To calc on criteria

        Args:
            criteria (str): Criteria
            status (dict): Hero status

        Returns:
            int: mark of the criteria
        """
        if criteria == 'SP':
            return status['SP']
        elif criteria == 'HP':
            return status['HP']
        elif criteria == 'HPS':
            return status['SP'] * status['HP']
        elif criteria == 'AT':
            vl = status['AT']
            if extra['IgnorePC'] == 0:
                if status['PC']:
                    vl *= 1.2
            return vl
        elif criteria == 'CT':
            return status['CT']
        elif criteria == 'CMG':
            vl = 100 - status['CT'] + status['CT'] * status['CD'] / 100
            if extra['IgnoreFU'] == 0:
                if status['FU']:
                    vl *= 1.3
            return vl
        elif criteria == 'CPS':
            vl = self.calc_criteria('CMG', extra, status)
            vl *= self.calc_criteria('SP', extra, status)
            return vl
        elif criteria == 'DMG':
            vl = self.calc_criteria('AT', extra, status)
            if extra['IgnoreCT'] == 0:
                vl *= self.calc_criteria('CMG', extra, status)
            return vl
        elif criteria == 'DPS':
            vl = self.calc_criteria('DMG', extra, status)
            vl *= self.calc_criteria('SP', extra, status)
            return vl
        elif criteria == 'DMGH':
            vl = self.calc_criteria('DMG', extra, status)
            # vl *= self.calc_criteria('HP', extra, status)
            return vl
        elif criteria == 'DPSS':
            vl = self.calc_criteria('DPS', extra, status)
            vl *= self.calc_criteria('SP', extra, status)
            return vl
        elif criteria == 'DPSD':
            vl = self.calc_criteria('DPS', extra, status)
            # vl *= self.calc_criteria('DF', extra, status)
            return vl
        elif criteria == 'DPSH':
            vl = self.calc_criteria('DPS', extra, status)
            # vl *= self.calc_criteria('HP', extra, status)
            return vl
        elif criteria == 'HT':
            return status['HT']
        elif criteria == 'EV':
            return status['EV']
        elif criteria == 'TANK':
            return status['HP'] * (1 + status['DF'] / 300)
        elif criteria == 'DF':
            vl = status['DF']
            return vl
        else:
            print('unexpected criteria {nm}'.format(nm=criteria))
def load_excel(field, workbook, holder):
    """A unified interface to load Excel data
    It can be used to load:
        Hero data
        Item data
        Artifact data

    Args:
        field (str): What kind of data is to be loaded
        workbook (Workbook): The workbook to be loaded
        holder (dict/list): The holder of extracted data
    """
    if field == 'hero':
        if 'hero' in workbook.sheetnames:
            load_sheet_hero(workbook['hero'], holder)
    elif field == 'plan':
        for ws in workbook:
            if ws.title in holder:
                print('error duplicated plan name')
                continue
            else:
                holder[ws.title] = []
            load_sheet_plan(ws, holder[ws.title])
    elif field == 'artifact':
        load_sheet_artifact(workbook[field], holder)
    else:
        load_sheet_item(workbook[field], holder[field])
def load_sheet_artifact(worksheet, holder):
    """Load artifact data from Excel

    Args:
        worksheet (Worksheet): The worksheet to be loaded
        holder (dict): The holder of extracted data
    """
    for rw in worksheet.iter_rows(min_row=2):
        name = rw[0].value
        ATa = rw[1].value
        HPa = rw[2].value
        # TBD
        # Extra attributes from artifact is not yet loaded
        if name in holder:
            print('error duplicated artifact')
            continue
        holder[name] = {'ATa': ATa, 'HPa': HPa}
def load_sheet_item(worksheet, holder):
    """Load item data from Excel

    Args:
        worksheet (Worksheet): The worksheet to be loaded
        holder (dict): The holder of extracted data
    """
    idx = 0
    for rw in worksheet.rows:
        if rw[0].value == None:
            continue
        itm = {}
        # Index
        itm['id'] = idx
        itm['set'] = rw[0].value
        itm['attributes'] = []
        for i in range(5):
            if rw[i * 2 + 1].value == None:
                continue
            itm['attributes'].append(
                {'type': rw[i * 2 + 1].value, 'value': rw[i * 2 + 2].value}
            )
        holder.append(itm)
        # Increase index
        idx += 1
def load_sheet_hero(worksheet, holder):
    """Load hero data from Excel

    Args:
        worksheet (Worksheet): The worksheet to be loaded
        holder (dict): The holder of extracted data
    """
    for rw in worksheet.iter_rows(min_row=2):
        # Check valid line
        if rw[0].value is None:
            continue
        hero = Sample()
        hero.load({
            'AT': rw[1].value,
            'DF': rw[2].value,
            'HP': rw[3].value,
            'SP': rw[4].value,
            'CT': rw[5].value,
            'CD': rw[6].value,
            'HT': rw[7].value,
            'EV': rw[8].value,
            'lv': rw[9].value,
            'rank': rw[10].value,
            'star': rw[11].value,
        })
        holder[rw[0].value] = hero
def load_sheet_plan(worksheet, holder):
    """Load plan data from Excel

    Args:
        worksheet (Worksheet): The worksheet to be loaded
        holder (dict): The holder of extracted data
    """
    for rw in worksheet.iter_rows(min_row=2):
        # Check valid line
        if rw[0].value is None:
            continue
        data = {}
        data['name'] = rw[0].value
        # Formula
        formula = []
        for i in range(1, 3):
            if rw[i].value != None:
                formula.append(rw[i].value)
        data['formula'] = formula
        # Thresholds
        data['thresholds'] = {'states': {}}
        # - Sets
        data['thresholds']['sets'] = set()
        for idx, field in PLAN_COL_THRESHOLD_SET.items():
            if rw[idx].value in ('x', '1', 1):
                data['thresholds']['sets'].add(field)
        # - States
        for idx, field in PLAN_COL_THRESHOLD_STATE.items():
            if rw[idx].value != None:
                value = rw[idx].value
                value_min = None
                value_max = None
                try:
                    if ';' in value:
                        value_min = float(value.split(';')[0])
                        value_max = float(value.split(';')[1])
                    else:
                        value_min = float(value)
                except TypeError:
                    value_min = value
                if value_min != None:
                    data['thresholds']['states'][field+'o'] = value_min
                if value_max != None:
                    data['thresholds']['states'][field+'u'] = value_max
        # States
        data['states'] = {}
        for idx, field in PLAN_COL_STATE.items():
            if rw[idx].value == None:
                data['states'][field] = 0
            else:
                data['states'][field] = rw[idx].value
        # Artifact
        data['artifact'] = rw[PLAN_COL_ARTIFACT].value
        # Extra
        data['extra'] = {}
        for idx, field in PLAN_COL_EXTRA.items():
            if rw[idx].value == None:
                data['extra'][field] = 0
            else:
                data['extra'][field] = rw[idx].value
        holder.append(data)
def calc_benchmark_group(items, hero, formula, thresholds, idx, total, queue, priority, flg_debug=False):
    """Calculate the best benchmark of a group data
    It shall do several calculation with different comibination of items sets
    and find out the best one.

    Args:
        items (dict): Item data which is prioritized
        hero (Hero): Target hero
        idx (int): Multi-thread thread index
        total (int): Multi-thread thread number
        queue (Queue): Holder for returned data
        priority (int): Iteration depth for iteration algorithm
        flg_debug (bool, optional): Debug flag. Defaults to False.
    """
    if flg_debug:
        print('calc_benchmark_group called with')
        print('thread {idx} of {total}'.format(idx=idx, total=total))
        print('priority={priority}'.format(priority=priority))
    result = None
    # 1. Use new ones only
    if priority < min([
        len(items['weapon']),
        len(items['head']),
        len(items['armor']),
        len(items['neck']),
        len(items['ring']),
        len(items['shoe']),
    ]):
        # Init data holder
        items_flattened = {}
        for il in ITEMS_LIST:
            items_flattened[il[0][0]] = []
        # Loop item kinds
        for il in ITEMS_LIST:
            # Loop items in target priority
            for itm in items[il[0]][priority]:
                # Add into holder
                items_flattened[il[0][0]].append(itm['item'])
        if flg_debug:
            print('Thread {nm} joins calculating A {pos}'.format(
                nm=idx,
                pos=len(items_flattened['w'])*len(items_flattened['h'])*len(items_flattened['a'])*len(items_flattened['n'])*len(items_flattened['r']*len(items_flattened['s']))
            ))
        # Put all items in target priority to calculate
        result = calc_benchmark_group_on_items_set(
            hero, items_flattened, formula, thresholds, total, idx, result, flg_debug=flg_debug
        )
    # 2. Use new ones to combine old ones
    # Loop all kinds of items
    for il in ITEMS_LIST:
        # New data exists?
        if priority < len(items[il[0]]):
            # Init data holder
            items_flattened = {}
            for il_1 in ITEMS_LIST:
                items_flattened[il_1[0][0]] = []
            # Add the target item kind
            for itm in items[il[0]][priority]:
                items_flattened[il[0][0]].append(itm['item'])
            # Use the rest item kinds with all previous priority
            for i in range(priority):
                for il_1 in ITEMS_LIST:
                    # Escape target kind
                    if il[0] == il_1[0]:
                        continue
                    # Prevent exceed list
                    if i < len(items[il_1[0]]):
                        for itm in items[il_1[0]][i]:
                            items_flattened[il_1[0][0]].append(itm['item'])
            if flg_debug:
                print('Thread {nm} joins calculating B {pos}'.format(
                    nm=idx,
                    pos=len(items_flattened['w'])*len(items_flattened['h'])*len(items_flattened['a'])*len(items_flattened['n'])*len(items_flattened['r']*len(items_flattened['s']))
                ))
            result = calc_benchmark_group_on_items_set(
                hero, items_flattened, formula, thresholds, total, idx, result, flg_debug=flg_debug
            )
    queue.put(result)
def calc_benchmark_group_on_items_set(
    hero, items, formula, thresholds, thread_num, idx_thread, result_prev=None, flg_debug=False
):
    """Calculate the best benchmark with an items set

    Args:
        hero (Hero): Target hero
        items (dict): Items set, already de-priorized
        thresholds (dict): Thresholds criteria
        thread_num (int): Thread number
        idx_thread (int): Thread index
        result_prev (dict, optional): The last calculated result. Defaults to
            None.
        flg_debug (bool, optional): Debug flag. Defaults to False.

    Returns:
        dict: The calculated best result
    """
    # By default benchmark parameter number is 3
    if result_prev == None:
        benchmark_best = [0, 0, 0]
        set_best = None
        hero_st = None
    else:
        benchmark_best = result_prev['benchmark_best']
        set_best = result_prev['set_best']
        hero_st = result_prev['hero_st']
    # An index for multi processing
    idx_thread_current = thread_num - 1
    for wp in items['w']:
        for hd in items['h']:
            for am in items['a']:
                for nk in items['n']:
                    for rg in items['r']:
                        for sh in items['s']:
                            # Filter for multi processing
                            # Loop index for multi processing
                            if idx_thread_current == thread_num - 1:
                                idx_thread_current = 0
                            else:
                                idx_thread_current += 1
                            if idx_thread == idx_thread_current:
                                hero.strip()
                                hero.equip(wp)
                                hero.equip(hd)
                                hero.equip(am)
                                hero.equip(nk)
                                hero.equip(rg)
                                hero.equip(sh)
                                # Check sets criteria
                                flg_chk_sets = True
                                sets = hero.get_sets()
                                for i in thresholds['sets']:
                                    if i not in sets:
                                        flg_chk_sets = False
                                if not flg_chk_sets:
                                    continue
                                # Update status
                                status = hero.get_status()
                                # Check thresholds pass?
                                if hero.chk_thresholds(thresholds['states'], status=status):
                                    benchmark = hero.get_benchmark(formula, extra, status=status)
                                    # Check benchmark valid
                                    if benchmark == None:
                                        continue
                                    # Loop into all parameters
                                    for i in range(len(benchmark)):
                                        # New is better
                                        if benchmark_best[i] < benchmark[i]:
                                            # Use new
                                            benchmark_best = benchmark
                                            if 'id' not in wp:
                                                print('error')
                                            set_best = [
                                                wp['id'],
                                                hd['id'],
                                                am['id'],
                                                nk['id'],
                                                rg['id'],
                                                sh['id'],
                                            ]
                                            hero_st = {}
                                            for bsl in BASIC_STATES_LIST:
                                                hero_st[bsl] = status[bsl]
                                        # New is worse
                                        elif benchmark_best[i] > benchmark[i]:
                                            # Next please
                                            break
                                        # New is same
                                        else:
                                            # Go to next parameter
                                            pass
    return {
        'benchmark_best': benchmark_best,
        'set_best': set_best,
        'hero_st': hero_st
    }
def calc_item_score_on_formula(item, formula, extra):
    """Calculate score for the item according to the formula
    The result shall be a list containing powers from each criteria

    Args:
        item (dict): Item data
        formula (list): Formula list

    Returns:
        [list]: each element is the score for one criteria
    """
    result = []
    # Loop all criteria
    for criteria in formula:
        # Add the power to result
        result.append(calc_item_score_on_criteria(item, criteria, extra))
    return result
def calc_item_score_on_criteria(item, criteria, extra, result=None):
    """Calculate score for the item according to the criteria
    The result shall be a list containing the min and max power pairs
    Each element in the list shall be a list with min and max power
    The values can be regarded as min powers
    The sets can be regarded as max powers cause it may improve result
    only in some of the cases

    Args:
        item (dict): item data
        criteria (str): criteria used
        result (list/None, optional): The latest result

    Returns:
        list: result containing each criteria score
    """
    if result == None:
        result = []
    if criteria == 'AT':
        # AT
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'ATa':
                value = attribute['value']
        result.append([value, value])
        value = 0
        for attribute in item['attributes']:
           if attribute['type'] == 'ATp':
                value = attribute['value']
        if item['set'] == 'AT':
            result.append([value, value + 35])
        else:
            result.append([value, value])
    elif criteria == 'CT':
        # CT
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'CTa':
                value = attribute['value']
        if item['set'] == 'CT':
            result.append([value, value + 12])
        else:
            result.append([value, value])
    elif criteria == 'CD':
        # CD
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'CDa':
                value = attribute['value']
        if item['set'] == 'CD':
            result.append([value, value + 40])
        else:
            result.append([value, value])
    elif criteria == 'SP':
        # SP
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'SPa':
                value = attribute['value']
        result.append([value, value])
        if item['set'] == 'SP':
            result.append([0, 1])
        else:
            result.append([0, 0])
    elif criteria == 'CMG':
        # CT + CD
        result = calc_item_score_on_criteria(item, 'CT', extra, result)
        result = calc_item_score_on_criteria(item, 'CD', extra, result)
    elif criteria == 'CPS':
        # CT + CD + SP = CMG + SP
        result = calc_item_score_on_criteria(item, 'DMG', extra, result)
        result = calc_item_score_on_criteria(item, 'SP', extra, result)
    elif criteria == 'DMG':
        # AT + (CT)
        result = calc_item_score_on_criteria(item, 'AT', extra, result)
        if extra['IgnoreCT'] == 0:
            result = calc_item_score_on_criteria(item, 'CMG', extra, result)
    elif criteria == 'DMGH':
        # DMG + HP
        result = calc_item_score_on_criteria(item, 'DMG', extra, result)
        # result = calc_item_score_on_criteria(item, 'HP', extra, result)
    elif criteria in ('DPS', 'DPSS'):
        # DMG + SP
        result = calc_item_score_on_criteria(item, 'DMG', extra, result)
        result = calc_item_score_on_criteria(item, 'SP', extra, result)
    elif criteria in ('DPSD', ):
        # DPS + DF
        result = calc_item_score_on_criteria(item, 'DPS', extra, result)
        # result = calc_item_score_on_criteria(item, 'DF', extra, result)
    elif criteria in ('DPSH', ):
        # DPS + HP
        result = calc_item_score_on_criteria(item, 'DPS', extra, result)
        # result = calc_item_score_on_criteria(item, 'HP', extra, result)
    elif criteria == 'HP':
        # HP
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'HPa':
                value = attribute['value']
        result.append([value, value])
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'HPp':
                value = attribute['value']
        if item['set'] == 'HP':
            result.append([value, value + 15])
        else:
            result.append([value, value])
    elif criteria == 'HPS':
        # HP + SP
        result = calc_item_score_on_criteria(item, 'HP', extra, result)
        result = calc_item_score_on_criteria(item, 'SP', extra, result)
    elif criteria == 'DF':
        # DF
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'DFa':
                value = attribute['value']
        result.append([value, value])
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'DFp':
                value = attribute['value']
        if item['set'] == 'DF':
            result.append([value, value + 15])
        else:
            result.append([value, value])
    elif criteria == 'TANK':
        # HP, DF
        result = calc_item_score_on_criteria(item, 'HP', extra, result)
        result = calc_item_score_on_criteria(item, 'DF', extra, result)
    elif criteria == 'HT':
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'HTa':
                value = attribute['value']
        # HT
        if item['set'] == 'HT':
            result.append([value, value + 20])
        else:
            result.append([value, value])
    elif criteria == 'EV':
        # EV
        value = 0
        for attribute in item['attributes']:
            if attribute['type'] == 'EVa':
                value = attribute['value']
        if item['set'] == 'EV':
            result.append([value, value + 20])
        else:
            result.append([value, value])
    else:
        print('unknown criteria {m}'.format(m=criteria))
    # Ignore FU
    if extra['IgnoreFU'] == 0:
        if item['set'] == 'FU':
            result.append([0, 1])
        else:
            result.append([0, 0])
    # Ignore PC
    if extra['IgnorePC'] == 0:
        if item['set'] == 'PC':
            result.append([0, 1])
        else:
            result.append([0, 0])
    return result
def compare_score(score_0, score_1):
    """Compare score
    If at least one pair compare is worse and no is better or not worse or
    unknown, regard as worse
    If at least one pair compare is better and no is worse or not better or
    unknown, regard as better
    Args:
        score_0 (list): compared score
        score_1 (list): comparing score

    Returns:
        str: 'better'/'worse'/'same'/'unknown'
    """
    flg_same = True
    idx = 0
    # Loop into all criteria scores
    for scs in score_0:
        flg_better = False
        flg_worse = False
        flg_not_better = False
        flg_not_worse = False
        flg_unknown = False
        idx_2 = 0
        for sc in scs:
            result = compare_min_max(sc, score_1[idx][idx_2])
            if result != 'same':
                flg_same = False
                if result == 'better':
                    flg_better = True
                elif result == 'worse':
                    flg_worse = True
                elif result == 'not better':
                    flg_not_better = True
                elif result == 'not worse':
                    flg_not_worse = True
                elif result == 'unknown':
                    flg_unknown = True
            idx_2 += 1
        # Better?
        if flg_better and not (flg_worse or flg_not_better or flg_unknown):
            return 'better'
        # Worse?
        if flg_worse and not (flg_better or flg_not_worse or flg_unknown):
            return 'worse'
        # Same?
        if flg_same:
            # Loop to next group
            idx += 1
        # Unknown? Regard as unknown and exit
        break
    # No difference found
    if flg_same:
        return 'same'
    # Otherwise unknown
    return 'unknown'
def compare_min_max(pair_0, pair_1):
    """Compare min max value pairs

    Args:
        pair_0 (list): min max value pairs comparing
        pair_1 (list): min max value pairs compared

    Returns:
        str: better/worse/not better/not worse/unknown
    """
    # Min value better
    if pair_0[0] > pair_1[0]:
        # Max value better
        if pair_0[1] > pair_1[1]:
            return 'better'
        # Max value same
        elif pair_0[1] == pair_1[1]:
            return 'not worse'
        # Max value worse
        else:
            return 'unknown'
    # Min value same
    elif pair_0[0] == pair_1[0]:
        # Max value better
        if pair_0[1] > pair_1[1]:
            return 'not worse'
        # Max value same
        elif pair_0[1] == pair_1[1]:
            return 'same'
        # Max value worse
        else:
            return 'not better'
    # Min value worse
    else:
        # Max value better
        if pair_0[1] > pair_1[1]:
            return 'unknown'
        # Max value same
        elif pair_0[1] == pair_1[1]:
            return 'not better'
        # Max value worse
        else:
            return 'worse'
def priorize_items_by_formula(items, formula, extra, debug=False):
    """Priorize items according to the formula.
    Each criteria is used to filter all the rest candidate items until all
    criteria is checked.
    Only targeted crteria can be counted, not thresholded ones
    The result shall be a list containing groups of items whoes power are equal
    Each group shall be marked with the power of each criteria as a list
    So that the best groups will be processed first to speed up
    If the best groups cannot provide a matched result (due to thresholds),
    Next group shall be used for further iteration

    Args:
        items ([list]): The original items pool
        formula ([list]): consists of criteria

    Returns:
        [dict]: priorized items in 6 classes
    """
    result = []
    # Loop all items in candidates
    for item in items:
        # Calculate this item
        score = calc_item_score_on_formula(item, formula, extra)
        if debug:
            print('-----------------------------------------')
            print('add item with score')
            print(score)
        # Find the matching group
        idx_results = 0
        flg_added = False
        # Loop into all groups from top to bottom
        for rslts in result:
            idx_results += 1
            # Loop all items in a group
            flg_stay = False
            not_beaten_results = []
            beaten_results = []
            idx_result = 0
            for rslt in rslts:
                # Compare
                compare_result = compare_score(score, rslt['score'])
                # New one is same
                if compare_result == 'same':
                    # Directly add into same group
                    result[idx_results - 1].append(
                        {'score': score, 'item': item}
                    )
                    flg_added = True
                    break
                # New one is unknown
                elif compare_result == 'unknown':
                    # Stay in this group
                    flg_stay = True
                    # Add compared as not beaten
                    not_beaten_results.append(rslt)
                # New one is better
                elif compare_result == 'better':
                    # Add compared as beaten
                    beaten_results.append(rslt)
                # New one is worse
                elif compare_result == 'worse':
                    # Loop to next group
                    break
                idx_result += 1
            else:
                # Stay in this group
                if flg_stay:
                    # Refresh this group with not beaten ones
                    result[idx_results - 1] = not_beaten_results
                    # Add into this group
                    result[idx_results - 1].append(
                        {'score': score, 'item': item}
                    )
                    flg_added = True
                    # Beat some of the candidate to a new lower group
                    if len(beaten_results) > 0:
                        result.insert(
                            idx_results, beaten_results
                        )
                # Not stay in this group, means nothing is unknown or same
                else:
                    # This one has beaten someone
                    if len(beaten_results) > 0:
                        # This one has been beaten
                        if len(not_beaten_results) > 0:
                            # This shall not happen
                            print('error')
                        else:
                            # Create a new group ahead
                            result.insert(
                                idx_results - 1, [{'score': score, 'item': item}]
                            )
                            flg_added = True
                    else:
                        # Loop to next group
                        pass                
            if flg_added:
                break
        else:
            result.append([{'score': score, 'item': item}])
        if debug:
            print('the result is now')
            idx = 0
            for rslts in result:
                print('rank {nm}'.format(nm=idx))
                for rslt in rslts:
                    print('{id}{score}'.format(id=rslt['item']['id'], score=rslt['score']))
                idx += 1
    return result
def get_data_prev(path):
    data = {}
    rp_find_num = re.compile('\d+')
    try:
        fh = open(path, 'r')
    except:
        return data
    step = 'hero'
    hero = ''
    for line in fh:
        if step == 'hero':
            hero = line[:-1]
            if hero.endswith('\t结果不唯一'):
                hero = hero[:-6]
            step = 'item'
        elif line.startswith('武器\t'):
            weapon_idx = rp_find_num.search(line).group(0)
            weapon_idx = int(weapon_idx) - 1
        elif line.startswith('头盔\t'):
            head_idx = rp_find_num.search(line).group(0)
            head_idx = int(head_idx) - 1
        elif line.startswith('盔甲\t'):
            armor_idx = rp_find_num.search(line).group(0)
            armor_idx = int(armor_idx) - 1
        elif line.startswith('项链\t'):
            neck_idx = rp_find_num.search(line).group(0)
            neck_idx = int(neck_idx) - 1
        elif line.startswith('戒指\t'):
            ring_idx = rp_find_num.search(line).group(0)
            ring_idx = int(ring_idx) - 1
        elif line.startswith('鞋子\t'):
            shoe_idx = rp_find_num.search(line).group(0)
            shoe_idx = int(shoe_idx) - 1
        elif line.startswith('速度'):
            step = 'hero'
            # Store current data
            data[hero] = {
                'weapon': weapon_idx,
                'head': head_idx,
                'armor': armor_idx,
                'neck': neck_idx,
                'ring': ring_idx,
                'shoe': shoe_idx,
            }
        else:
            pass
    # except:
    #     return None
    return data
if __name__ == '__main__':
    # Debug flag for info
    flg_debug = False
    # Debug flag to force single thread
    flg_single_thread = False
    # Prepare database
    D = Database()
    # Get config
    cfg = ConfigParser()
    cfg.read('config.ini')
    task_number = int(cfg['CompuPower']['ThreadNumber'])
    pth_plan = cfg['Paths']['Plans']
    pth_data = cfg['Paths']['Data']
    # Load items and artifacts data
    items = {}
    for il in ITEMS_LIST:
        items[il[0]] = []
    artifacts = {}
    for pth in os.listdir(pth_data):
        if pth.endswith('.xlsx') and not pth.startswith('~$'):
            workbook = load_workbook(os.path.join(pth_data, pth), data_only=True)
            for il in ITEMS_LIST:
                load_excel(il[0], workbook, items)
            load_excel('artifact', workbook, artifacts)
    # Load hero data
    heroes = {}
    for pth in os.listdir(pth_data):
        if pth.endswith('.xlsx') and not pth.startswith('~$'):
            workbook = load_workbook(os.path.join(pth_data, pth), data_only=True)
            load_excel('hero', workbook, heroes)
    for name, data in heroes.items():
        lv = data.lv
        rank = data.rank
        star = data.star
        hero = D.make_by_name(name, lv, rank, star)
        if hero == False:
            print(name)
            input()
        states = hero.get_states()
        if data.AT == None:
            data.AT = states['at']
            data.DF = states['df']
            data.HP = states['hp']
            data.SP = states['sp']
            data.CT = 100 * states['ct']
            data.CD = 100 * states['cd']
            data.HT = 100 * states['ht']
            data.EV = 100 * states['ev']
            continue
        if data.AT != states['at']:
            print(name)
            print('at')
            print(data.AT)
            print(states['at'])
            data.AT = states['at']
        if data.DF != states['df']:
            print(name)
            print('df')
            print(data.DF)
            print(states['df'])
            data.DF = states['df']
        if data.HP != states['hp']:
            print(name)
            print('hp')
            print(data.HP)
            print(states['hp'])
            data.HP = states['hp']
        if data.SP != states['sp']:
            print(name)
            print('sp')
            print(data.SP)
            print(states['sp'])
            data.SP = states['sp']
        if data.CT != 100 * states['ct']:
            print(name)
            print('ct')
            print(data.CT)
            print(states['ct'])
            data.CT = 100 * states['ct']
        if data.CD != 100 * states['cd']:
            print(name)
            print('cd')
            print(data.CD)
            print(states['cd'])
            data.CD = 100 * states['cd']
        # if data.TO != states['to']:
        #     input('to')
        if data.HT != 100 * states['ht']:
            print(name)
            print('ht')
            print(data.HT)
            print(states['ht'])
            data.HT = 100 * states['ht']
        if data.EV != 100 * states['ev']:
            print(name)
            print('ev')
            print(data.EV)
            print(states['ev'])
            data.EV = 100 * states['ev']
    # Load plans
    plans = {}
    for pth in os.listdir(pth_plan):
        if pth.endswith('.xlsx') and not pth.startswith('~$'):
            # Get name of plan set
            name = pth[:-5]
            if name in plans:
                print('error duplicated plan set name')
                continue
            plans[name] = {}
            workbook = load_workbook(os.path.join(pth_plan, pth), data_only=True)
            load_excel('plan', workbook, plans[name])
    # Step 3: Results
    # Create folder for results
    try:
        os.mkdir('Results')
    except:
        pass
    for plan_set_name in sorted(plans):
        print(plan_set_name)
        plan_set = plans[plan_set_name]
        # Create folder for plan set
        try:
            os.mkdir(os.path.join('Results', plan_set_name))
        except:
            pass
        for plan_name in sorted(plan_set):
            print(plan_name)
            plan = plan_set[plan_name]
            path_result = os.path.join('Results', plan_set_name, plan_name + '.txt')
            # Try to open existed file for continuous calculation
            data_prev = get_data_prev(path_result)
            if data_prev != None:
                f = open(path_result, 'a', encoding="utf-8")
            else:
                f = open(path_result, 'w', encoding="utf-8")
            f.close()
            # Copy items
            items_copy = items.copy()
            # Process previous data
            for hero_name, hero_data in data_prev.items():
                print(f'skip plan for [{hero_name}]')
                # Remove this hero plan
                for hero_plan in plan:
                    if hero_plan['name'] == hero_name:
                        plan.pop(plan.index(hero_plan))
                # Mark items used previously
                for item_type, item_idx in hero_data.items():
                    for item_data in items_copy[item_type]:
                        if item_data['id'] == item_idx:
                            items_copy[item_type].pop(items_copy[item_type].index(item_data))
            # Calculate this plan
            for hero_plan in plan:
                hero_name = hero_plan['name']
                # Info
                print(f"Let's make {hero_name} great again!")
                # Get formula
                formula = hero_plan['formula']
                # Get extra
                extra = hero_plan['extra']
                # Check formula valid
                if len(formula) == 0:
                    continue
                # Init flags
                flg_found_same_results = False
                if hero_name not in heroes:
                    print('hero cannot be found. Check output file for details')
                    f = open(path_result, 'a', encoding="utf-8")
                    f.write('error hero [{nm}] not found\n'.format(nm=hero_name))
                    f.close()
                    continue
                hero = heroes[hero_name]
                # Load plan to the hero
                (ec, info) = hero.load_plan(hero_plan)
                if ec != None:
                    f = open(path_result, 'a', encoding="utf-8")
                    f.write('error artifact [{nm}] not found\n'.format(nm=name))
                    f.close()
                # Get Thresholds
                thresholds = hero_plan['thresholds']
                # Priorize items
                items_priorized = {}
                # Loop all 6 classes to be filtered
                for itm_type, itms in items_copy.items():
                    items_priorized[itm_type] = priorize_items_by_formula(
                        itms, formula, extra
                    )
                max_loop = max(
                    len(items_priorized['weapon']),
                    len(items_priorized['head']),
                    len(items_priorized['armor']),
                    len(items_priorized['neck']),
                    len(items_priorized['ring']),
                    len(items_priorized['shoe']),
                )
                counter = [0, 0, 0, 0, 0, 0]
                for i in range(max_loop):
                    print('Start calculation round {id}'.format(id=i))
                    # Loop all kinds of items
                    for il in ITEMS_LIST:
                        # Increase the counter
                        if len(items_priorized[il[0]]) > i:
                            counter[il[1]] += len(items_priorized[il[0]][i])
                    buff = 'Using '
                    for il in ITEMS_LIST:
                        buff += '{num}/{total} '.format(num=counter[il[1]], total=len(items[il[0]]))
                    buff += 'items'
                    print(buff)
                    # Debug with single thread
                    if flg_single_thread:
                        hero = hero.copy()
                        q = mp.Queue()
                        calc_benchmark_group(
                            items_priorized, hero, formula, thresholds, 0, 1, q, i, flg_debug
                        )
                        result_best = q.get()
                        if result_best['set_best'] == None:
                            result_best = None
                    # Normal running with multiple threads
                    else:
                        processes = []
                        # Queue for results
                        q = mp.Queue()
                        for idx_task in range(task_number):
                            hero = hero.copy()
                            process = mp.Process(
                                target=calc_benchmark_group,
                                args=(
                                    items_priorized,
                                    hero,
                                    formula,
                                    thresholds,
                                    idx_task,
                                    task_number,
                                    q,
                                    i,
                                    flg_debug
                                )
                            )
                            processes.append(process)
                            process.start()
                        for idx_task in range(task_number):
                            processes[idx_task].join()
                        # Choose best result
                        result_best = None
                        for idx_task in range(task_number):
                            result = q.get()
                            if result['set_best'] != None:
                                for j in range(len(result['benchmark_best'])):
                                    # New is better
                                    if result_best == None or (
                                        result['benchmark_best'][j] > result_best['benchmark_best'][j]
                                    ):
                                        # Update best result
                                        result_best = result
                                        # Stop comparing
                                        break
                                    # New is worse
                                    elif result['benchmark_best'][j] < result_best['benchmark_best'][j]:
                                        # Give up this result
                                        break
                                    # New is same
                                    else:
                                        # Compare next parameter
                                        pass
                                else:
                                    flg_found_same_results = True
                    # Qualified result found
                    if result_best != None:
                        with open(path_result, 'a', encoding="utf-8") as f:
                            f.write(hero_name)
                            if flg_found_same_results:
                                f.write('\t结果不唯一')
                            f.write('\n')
                            for il in ITEMS_LIST:
                                item = None
                                for itm in items_copy[il[0]]:
                                    if itm['id'] == result_best['set_best'][il[1]]:
                                        item = items_copy[il[0]].pop(items_copy[il[0]].index(itm))
                                        break
                                if item == None:
                                    print('error something is wrong')
                                f.write(PRINT_ITEMS_KINDS[il[0]])
                                f.write('\t第')
                                f.write(str(item['id'] + 1))
                                f.write('件\n')
                                f.write('\t')
                                f.write(' '.join(['套装', item['set']]))
                                f.write('\n')
                                for atr in item['attributes']:
                                    f.write('\t'.join(['', atr['type'], str(atr['value'])]))
                                f.write('\n')
                            f.write('攻击力\t{nm}\n'.format(nm=result_best['hero_st']['AT']))
                            f.write('防御力\t{nm}\n'.format(nm=result_best['hero_st']['DF']))
                            f.write('生命力\t{nm}\n'.format(nm=result_best['hero_st']['HP']))
                            f.write('暴击率\t{nm}\n'.format(nm=result_best['hero_st']['CT']))
                            f.write('暴击伤害\t{nm}\n'.format(nm=result_best['hero_st']['CD']))
                            f.write('效果命中\t{nm}\n'.format(nm=result_best['hero_st']['HT']))
                            f.write('效果抗性\t{nm}\n'.format(nm=result_best['hero_st']['EV']))
                            f.write('速度\t{nm}\n'.format(nm=result_best['hero_st']['SP']))
                        break
                else:
                    print('error no suits can be found')
